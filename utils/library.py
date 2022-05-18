from itertools import islice
import configparser
from openpyxl import load_workbook

    # Fetch Configurable parameters
config = configparser.ConfigParser()
config.read('Configuration.cfg')
lastCommonDataColumnNumber = int(config['Files']['lastCommonDataColumnNumber'])
lastRequestColumnNumber = int(config['Files']['lastRequestColumnNumber'])
expectedToColumn = int(config['Files']['expectedToColumn'])

class Common:

    # constructor for load workbook, workbook object & sheet object should be global since it will be used everywhere
    def __init__(self, FileNamePath, SheetName):
        global wb
        global sheet
        wb = load_workbook(FileNamePath)
        sheet = wb[SheetName]

    # Get Rowcount of a sheet
    def fetch_row_count(self):
        rows = sheet.max_row
        return rows

    # Get Rowcount of a sheet
    def fetch_column_count(self):
        col = sheet.max_column
        return col

    # insert takes 2 argumets , 1st argument is index , index starts with 0 and i is starting from 1,
    # hence we take index from i-1 i.e 1-1 i.e 0 , so insert(i-1 , cell.value)

    # Fetch column headings of xls sheet
    def fetch_key_names(self, fromCol, toCol):
        listing = []
        for i in range(fromCol, toCol):
            cell = sheet.cell(row=1, column=i)
            listing.insert(i - 1, cell.value)
        return listing

    # Convert the XLS row value into a dictionary object having key value pairs to create input request
    def return_dictionary(self, keyList, rowNumber):
        dictionary={}
        anotherDict={}

        for row in islice(sheet.values, 1, rowNumber):

            for i in range(0, lastCommonDataColumnNumber):
                dictionary[keyList[i]] = row[i]
            for i in range(lastCommonDataColumnNumber, lastRequestColumnNumber):
                if keyList[i] == "elementNames":
                    if row[i] is None:
                        anotherDict[keyList[i]] = row[i]
                    else:
                        elementNames_string = row[i]
                        elementNames_list = elementNames_string.split(",")
                        anotherDict[keyList[i]] = elementNames_list
                else:
                    anotherDict[keyList[i]] = row[i]

            dictionary['parameters'] = anotherDict
        return dictionary


    # To create dictionary for expected kay and values
    def return_expected_dictionary(self, expectedKeyList, rowNumber):
        expectedDict = {}
        for row in islice(sheet.values, 1, rowNumber):
            for i in range(lastRequestColumnNumber, expectedToColumn-1):
                expectedDict[expectedKeyList[i-lastRequestColumnNumber]] = row[i]
        return expectedDict